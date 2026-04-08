"""
Deal Desk Modeler — Single-file Streamlit app.
Run: streamlit run pricing_app.py
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date
from enum import Enum
from io import BytesIO
from pathlib import Path

import pandas as pd
import streamlit as st


# ══════════════════════════════════════════════════════════════════════════════
# Domain Types
# ══════════════════════════════════════════════════════════════════════════════


class PricingType(Enum):
    PER_USER = "Per User"
    FLAT_FEE = "Flat Fee"


class UpliftMode(Enum):
    MANUAL = "Manual Uplift"
    GOAL_SEEK = "Goal-Seek TCV"


@dataclass
class Product:
    name: str
    pricing_type: PricingType
    base_price: float


@dataclass
class Fee:
    name: str
    price: float


@dataclass
class YearResult:
    year: int
    uplift_pct: float
    users: int
    product_rates: dict[str, float]
    arr: float
    one_time_charges: dict[str, float]
    total_billed: float


@dataclass
class ScenarioConfig:
    name: str = "Scenario"
    term: int = 3
    payment_label: str = "Upfront (0 Days)"
    mode: UpliftMode = UpliftMode.MANUAL
    added_fees: list[str] = field(default_factory=list)
    target_tcv: float = 250_000.0


# ══════════════════════════════════════════════════════════════════════════════
# Constants
# ══════════════════════════════════════════════════════════════════════════════

MAX_TERM = 5
PAY_TERMS: dict[str, int] = {
    "Upfront (0 Days)": 0,
    "Net 30": 30,
    "Net 60": 60,
    "Net 90": 90,
}
UPLIFT_MIN, UPLIFT_MAX = 0, 30
GS_ITERATIONS = 50
GS_TOLERANCE = 1.0
GS_LOWER_BOUND = -0.50
GS_UPPER_BOUND = 2.0

DEFAULT_PRODUCTS: list[Product] = []
DEFAULT_FEES = [Fee("Custom Integration / Migration", 0.0)]

# ── Product Catalog (from Service Offerings List) ──
PRODUCT_CATALOG: list[dict] = [
    {"name": "Additional Custom Reports (Support)", "list_price": 275.0, "cadence": "One-Time"},
    {"name": "Additional Custom Reports (Professional Services)", "list_price": 275.0, "cadence": "One-Time"},
    {"name": "Additional Languages Enhancement Grants Essential/Benevity Grants", "list_price": 5500.0, "cadence": "One-Time"},
    {"name": "Additional Languages Enhancement Grants Advanced/Versaic", "list_price": 13200.0, "cadence": "One-Time"},
    {"name": "Additional Localized Experience - Enhancement - Implementation Fee", "list_price": 7000.0, "cadence": "One-Time"},
    {"name": "Advanced Payroll - Enhancement - Implementation Fee", "list_price": 19320.0, "cadence": "One-Time"},
    {"name": "Advanced Workflow", "list_price": 275.0, "cadence": "One-Time"},
    {"name": "Adverse Media Reports", "list_price": 5000.0, "cadence": "Recurring"},
    {"name": "Basic Payroll Giving - Enhancement - Implementation Fee", "list_price": 9000.0, "cadence": "One-Time"},
    {"name": "Basic Workflow (up to 5 stages)", "list_price": 6600.0, "cadence": "One-Time"},
    {"name": "Challenges Implementation", "list_price": 5000.0, "cadence": "One-Time"},
    {"name": "Client Onsite", "list_price": 275.0, "cadence": "One-Time"},
    {"name": "Core Implementation Service: Employee Engagement", "list_price": 0.0, "cadence": "One-Time"},
    {"name": "Custom Budget Tags", "list_price": 8250.0, "cadence": "One-Time"},
    {"name": "Custom Fraud Rules/Client Giving Match Protection (Review-Based)", "list_price": 20000.0, "cadence": "Recurring"},
    {"name": "Custom Historical Data Import", "list_price": 275.0, "cadence": "One-Time"},
    {"name": "Custom Managed Services (One-Time)", "list_price": 300.0, "cadence": "One-Time"},
    {"name": "Custom Managed Services (Recurring)", "list_price": 300.0, "cadence": "Recurring"},
    {"name": "Custom Reports", "list_price": 0.0, "cadence": "Recurring"},
    {"name": "Custom Reports Enhancement", "list_price": 0.0, "cadence": "One-Time"},
    {"name": "Data Export Implementation MFT", "list_price": 5000.0, "cadence": "One-Time"},
    {"name": "Data Transformations - Enhancement - Implementation Fee", "list_price": 2750.0, "cadence": "One-Time"},
    {"name": "Dedicated Phone Line (End User)", "list_price": 0.0, "cadence": "Recurring"},
    {"name": "Donation Report Uploads", "list_price": 5000.0, "cadence": "Recurring"},
    {"name": "Employee Engagement URL Change", "list_price": 11000.0, "cadence": "One-Time"},
    {"name": "Employee Groups \u2013 Budget Enhancement", "list_price": 4000.0, "cadence": "One-Time"},
    {"name": "Employee Groups \u2013 Chapters Advanced Enhancement", "list_price": 8000.0, "cadence": "One-Time"},
    {"name": "Employee Groups \u2013 Chapters Essential Enhancement", "list_price": 4000.0, "cadence": "One-Time"},
    {"name": "Employee Groups \u2013 SSO Enhancement - Implementation Enhancement", "list_price": 2254.0, "cadence": "One-Time"},
    {"name": "eSignature or additional eSignatures", "list_price": 4950.0, "cadence": "One-Time"},
    {"name": "External Match Request Reviews", "list_price": 0.0, "cadence": "Recurring"},
    {"name": "External Portal (Up To 2 Tasks) (Inlc. Submitter Portal)", "list_price": 9625.0, "cadence": "One-Time"},
    {"name": "External Portal Enhancement", "list_price": 0.0, "cadence": "One-Time"},
    {"name": "Flat File Payment Integration Versaic - Setup", "list_price": 0.0, "cadence": "One-Time"},
    {"name": "GA: Supplementary Forms Enhancement", "list_price": 8250.0, "cadence": "One-Time"},
    {"name": "GE: Supplementary Forms Enhancement", "list_price": 6050.0, "cadence": "One-Time"},
    {"name": "Gift of Securities", "list_price": 180000.0, "cadence": "Recurring"},
    {"name": "Giving Implementation Fee", "list_price": 19000.0, "cadence": "One-Time"},
    {"name": "Giving Opportunity Reviews", "list_price": 0.0, "cadence": "Recurring"},
    {"name": "Global DAF", "list_price": 40000.0, "cadence": "Recurring"},
    {"name": "Global Program Bundle Implementation", "list_price": 19000.0, "cadence": "One-Time"},
    {"name": "Global Program Bundle Implementation (Existing)", "list_price": 15950.0, "cadence": "One-Time"},
    {"name": "Grants - Advanced Program + Workflow (Grants Essential)", "list_price": 4950.0, "cadence": "One-Time"},
    {"name": "Grants - Advanced Program + Workflow", "list_price": 14437.5, "cadence": "One-Time"},
    {"name": "Grants - Essential Program + Workflow", "list_price": 61325.0, "cadence": "One-Time"},
    {"name": "Grants - Essential Program + Workflow (Grants Essential)", "list_price": 32037.5, "cadence": "One-Time"},
    {"name": "Grants - External Portal Bundle", "list_price": 8937.5, "cadence": "One-Time"},
    {"name": "Grants Application Review Service", "list_price": 250.0, "cadence": "Recurring"},
    {"name": "Grants Management - Implementation Package - Core", "list_price": 35725.0, "cadence": "One-Time"},
    {"name": "Grants Management Plus Implementation", "list_price": 72540.0, "cadence": "One-Time"},
    {"name": "Grants Management Premium Implementation", "list_price": 0.0, "cadence": "One-Time"},
    {"name": "Grants Program Change (Minor, Medium, Complex)", "list_price": 275.0, "cadence": "One-Time"},
    {"name": "Historical Data Import", "list_price": 0.0, "cadence": "One-Time"},
    {"name": "HRIS Updates and/or Provider Change", "list_price": 8250.0, "cadence": "One-Time"},
    {"name": "Implementation Consultation", "list_price": 275.0, "cadence": "One-Time"},
    {"name": "Integrations Enhancement", "list_price": 275.0, "cadence": "One-Time"},
    {"name": "Local Login Authentication Enhancements", "list_price": 2000.0, "cadence": "One-Time"},
    {"name": "Merger or Acquisition", "list_price": 13750.0, "cadence": "One-Time"},
    {"name": "Modified Donation Report", "list_price": 5750.0, "cadence": "Recurring"},
    {"name": "Nonprofit Curation", "list_price": 1500.0, "cadence": "One-Time"},
    {"name": "Nonprofit Mission & Operations Screening", "list_price": 1600.0, "cadence": "Recurring"},
    {"name": "Off Cycle Disbursement", "list_price": 2500.0, "cadence": "One-Time"},
    {"name": "Organization and Key Personnel Screening", "list_price": 875.0, "cadence": "Recurring"},
    {"name": "Plus Implementation Service: Employee Engagement", "list_price": 0.0, "cadence": "One-Time"},
    {"name": "Plus Implementation Service: Employee Groups", "list_price": 0.0, "cadence": "One-Time"},
    {"name": "Post Approval Activities (Incl. Impact Report)", "list_price": 8250.0, "cadence": "One-Time"},
    {"name": "Premium Implementation Service: Employee Engagement", "list_price": 0.0, "cadence": "One-Time"},
    {"name": "Premium Implementation Service: Employee Groups", "list_price": 0.0, "cadence": "One-Time"},
    {"name": "Premium Nonprofit Screening", "list_price": 3200.0, "cadence": "Recurring"},
    {"name": "Premium Service Level Agreement (Client Admin)", "list_price": 0.0, "cadence": "Recurring"},
    {"name": "Premium Service Level Agreement (End User)", "list_price": 0.0, "cadence": "Recurring"},
    {"name": "Primary Forms Enhancement (All types)", "list_price": 8250.0, "cadence": "One-Time"},
    {"name": "Project Management - Enhancement - Implementation Fee", "list_price": 275.0, "cadence": "One-Time"},
    {"name": "Reversal", "list_price": 250.0, "cadence": "One-Time"},
    {"name": "Secular Process Review", "list_price": 62.5, "cadence": "Recurring"},
    {"name": "Service Reserve Units", "list_price": 500.0, "cadence": "One-Time"},
    {"name": "Site Consolidation", "list_price": 21000.0, "cadence": "One-Time"},
    {"name": "SSO Enhancement", "list_price": 2250.0, "cadence": "One-Time"},
    {"name": "Technical Account Manager (TAM)", "list_price": 0.0, "cadence": "Recurring"},
    {"name": "Technical Audit", "list_price": 0.0, "cadence": "One-Time"},
    {"name": "Technical Consultation - Grants Management", "list_price": 275.0, "cadence": "One-Time"},
    {"name": "Technical Program Consultation", "list_price": 12075.0, "cadence": "One-Time"},
    {"name": "User Data File Enhancements", "list_price": 275.0, "cadence": "One-Time"},
    {"name": "Volunteer Opportunity Reviews", "list_price": 0.0, "cadence": "Recurring"},
    {"name": "Volunteer Rewards Implementation Fee", "list_price": 9625.0, "cadence": "One-Time"},
    {"name": "Volunteer Time Reviews", "list_price": 0.0, "cadence": "Recurring"},
    {"name": "Volunteering Implementation Fee", "list_price": 9625.0, "cadence": "One-Time"},
    {"name": "Workday Integration Implementation Service", "list_price": 5200.0, "cadence": "One-Time"},
    {"name": "Workday Integration - Payroll Only", "list_price": 9000.0, "cadence": "One-Time"},
    {"name": "Workday Integration - User Data and Payroll", "list_price": 22000.0, "cadence": "One-Time"},
    {"name": "Workday Integration - User Data Only", "list_price": 13000.0, "cadence": "One-Time"},
]
CATALOG_LOOKUP: dict[str, dict] = {p["name"]: p for p in PRODUCT_CATALOG}
CATALOG_NAMES: list[str] = [p["name"] for p in PRODUCT_CATALOG]
DEFAULT_USERS = 0

SCENARIO_DEFAULTS: list[ScenarioConfig] = [
    ScenarioConfig("1 Yr Baseline", term=1),
    ScenarioConfig("3 YR - Deal", term=3),
    ScenarioConfig("Scenario 3", term=3),
    ScenarioConfig("Scenario 4", term=3),
]


# ══════════════════════════════════════════════════════════════════════════════
# Engine (pure functions — no Streamlit)
# ══════════════════════════════════════════════════════════════════════════════


def pad_to_term(values: list[float], length: int, fill: float) -> list[float]:
    """Extend `values` to `length` by repeating the last element, or `fill` if empty."""
    if not values:
        return [fill] * length
    out = list(values)
    while len(out) < length:
        out.append(out[-1])
    return out[:length]


def empty_discount_ledger() -> pd.DataFrame:
    return pd.DataFrame({
        "Year": pd.Series(dtype="int"),
        "Product": pd.Series(dtype="str"),
        "Discount (%)": pd.Series(dtype="float"),
    })


def _get_discount(ledger: pd.DataFrame, year: int, product_name: str) -> float:
    """Return the fractional discount (0-1) for a product in a given year."""
    if ledger.empty:
        return 0.0
    match = ledger[(ledger["Year"] == year) & (ledger["Product"] == product_name)]
    if match.empty:
        return 0.0
    return float(match["Discount (%)"].iloc[0]) / 100


def compute_effective_prices(
    products: list[Product],
    uplifts: list[float],
    ledger: pd.DataFrame,
    num_years: int,
) -> dict[str, list[float]]:
    """
    Build a year-indexed price schedule for each product.
    Single forward pass per product — O(products × years).
    """
    prices: dict[str, list[float]] = {}
    for product in products:
        yearly = []
        price = product.base_price
        for year in range(num_years):
            price *= 1 + uplifts[year]
            discount = _get_discount(ledger, year + 1, product.name)
            effective = price * (1 - discount)
            yearly.append(effective)
        prices[product.name] = yearly
    return prices


def run_model(
    users: list[int],
    products: list[Product],
    fees: list[Fee],
    added_fee_names: list[str],
    uplifts: list[float],
    payment_days: int,
    ledger: pd.DataFrame,
    num_years: int = MAX_TERM,
) -> list[YearResult]:
    uplifts = pad_to_term(uplifts, num_years, fill=0.0)
    price_schedule = compute_effective_prices(products, uplifts, ledger, num_years)

    results: list[YearResult] = []
    for year_idx in range(num_years):
        year_num = year_idx + 1
        user_count = users[year_idx]

        product_rates: dict[str, float] = {}
        arr = 0.0
        for product in products:
            rate = price_schedule[product.name][year_idx]
            product_rates[product.name] = rate
            if product.pricing_type == PricingType.PER_USER:
                arr += user_count * rate
            else:
                arr += rate

        one_time: dict[str, float] = {}
        ot_total = 0.0
        for fee in fees:
            key = f"{fee.name} (One-Time)"
            if year_num == 1 and fee.name in added_fee_names:
                one_time[key] = fee.price
                ot_total += fee.price
            else:
                one_time[key] = 0.0

        total_billed = arr + ot_total

        results.append(YearResult(
            year=year_num,
            uplift_pct=uplifts[year_idx],
            users=user_count,
            product_rates=product_rates,
            arr=arr,
            one_time_charges=one_time,
            total_billed=total_billed,
        ))
    return results


def goal_seek_uplift(
    target_tcv: float,
    users: list[int],
    products: list[Product],
    fees: list[Fee],
    added_fee_names: list[str],
    payment_days: int,
    ledger: pd.DataFrame,
    term: int,
) -> float:
    """Binary search for a flat uplift rate that hits `target_tcv`."""
    lo, hi = GS_LOWER_BOUND, GS_UPPER_BOUND
    best = 0.0
    for _ in range(GS_ITERATIONS):
        mid = (lo + hi) / 2
        schedule = run_model(
            users=users, products=products, fees=fees,
            added_fee_names=added_fee_names, uplifts=[mid] * term,
            payment_days=payment_days, ledger=ledger, num_years=term,
        )
        tcv = sum(r.total_billed for r in schedule)
        if tcv > target_tcv:
            hi = mid
        else:
            lo = mid
        best = mid
        if abs(tcv - target_tcv) < GS_TOLERANCE:
            break
    return best


def results_to_dataframe(results: list[YearResult]) -> pd.DataFrame:
    """Pivot YearResults: rows = metrics/products, columns = Year 1..N + Total."""
    if not results:
        return pd.DataFrame()

    year_cols = [f"Year {r.year}" for r in results]
    rows: list[dict[str, object]] = []

    # Users row
    user_row: dict[str, object] = {"": "Users"}
    for r in results:
        user_row[f"Year {r.year}"] = r.users
    user_row["Total"] = ""
    rows.append(user_row)

    # Uplift row
    uplift_row: dict[str, object] = {"": "Uplift"}
    for r in results:
        uplift_row[f"Year {r.year}"] = f"{r.uplift_pct * 100:.2f}%"
    uplift_row["Total"] = ""
    rows.append(uplift_row)

    # Product rate rows (recurring)
    if results[0].product_rates:
        for name in results[0].product_rates:
            rate_row: dict[str, object] = {"": f"{name} (Rate)"}
            for r in results:
                rate_row[f"Year {r.year}"] = r.product_rates.get(name, 0)
            rate_row["Total"] = ""
            rows.append(rate_row)

    # ARR row
    arr_row: dict[str, object] = {"": "ARR"}
    arr_total = 0.0
    for r in results:
        arr_row[f"Year {r.year}"] = r.arr
        arr_total += r.arr
    arr_row["Total"] = arr_total
    rows.append(arr_row)

    # One-time charge rows
    if results[0].one_time_charges:
        for key in results[0].one_time_charges:
            ot_row: dict[str, object] = {"": key}
            ot_total = 0.0
            for r in results:
                val = r.one_time_charges.get(key, 0)
                ot_row[f"Year {r.year}"] = val
                ot_total += val
            ot_row["Total"] = ot_total
            rows.append(ot_row)

    # Total Billed row
    billed_row: dict[str, object] = {"": "Total Billed"}
    billed_total = 0.0
    for r in results:
        billed_row[f"Year {r.year}"] = r.total_billed
        billed_total += r.total_billed
    billed_row["Total"] = billed_total
    rows.append(billed_row)

    return pd.DataFrame(rows)


def results_to_export_dataframe(label: str, results: list[YearResult]) -> pd.DataFrame:
    """
    Flatten YearResult list into an export-ready DataFrame for Google Sheets.
    Columns: Scenario, Year, Uplift (%), Users, one column per product rate, ARR.
    """
    rows = []
    for r in results:
        row: dict[str, object] = {
            "Scenario": label,
            "Year": r.year,
            "Uplift (%)": round(r.uplift_pct * 100, 4),
            "Users": r.users,
        }
        for name, rate in r.product_rates.items():
            row[f"{name} Rate ($)"] = round(rate, 4)
        row["ARR ($)"] = round(r.arr, 2)
        rows.append(row)
    return pd.DataFrame(rows)


def build_export_xlsx(results: list[dict]) -> bytes:
    """
    Build an XLSX with three tables:
      1. Scenario 1 schedule
      2. Scenario 2 schedule
      3. Billed Savings (Nominal) comparison
    Gray headers, bold totals, dollar/percent formatting.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill("solid", fgColor="F1F5F9")
    TOTAL_FILL  = PatternFill("solid", fgColor="E2E8F0")
    HEADER_FONT = Font(bold=True, size=10)
    TOTAL_FONT  = Font(bold=True, size=10)
    MONEY_FMT   = '"$"#,##0.00'
    UPLIFT_FMT  = '0.00"%"'   # value is already ×100
    PCT_FMT     = '0.00%'     # value is decimal (0.05 = 5%)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Deal Desk Export"
    row = 1

    def write_header(cols: list[str]) -> None:
        nonlocal row
        for c, name in enumerate(cols, 1):
            cell = ws.cell(row=row, column=c, value=name)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        row += 1

    def apply_fmt(cell, col_name: str, is_total: bool = False) -> None:
        if is_total:
            cell.fill = TOTAL_FILL
            cell.font = TOTAL_FONT
        if "Rate ($)" in col_name or col_name in ("ARR ($)", "Saving ($)") or col_name.endswith(" ($)"):
            cell.number_format = MONEY_FMT
        elif col_name == "Uplift (%)":
            cell.number_format = UPLIFT_FMT
        elif col_name == "Saving (%)":
            cell.number_format = PCT_FMT

    # ── Scenario tables (pivoted: rows=metrics, cols=years) ──
    export_scenarios = [s for s in results if s["Schedule"]]
    for scenario in export_scenarios:
        schedule = scenario["Schedule"]
        if not schedule:
            continue

        # Title row
        cell = ws.cell(row=row, column=1, value=scenario["Label"])
        cell.font = Font(bold=True, size=12)
        row += 1

        year_labels = [f"Year {r.year}" for r in schedule]
        write_header([""] + year_labels + ["Total"])

        # Collect metric rows
        metric_rows: list[tuple[str, list, object]] = []

        # Users
        metric_rows.append(("Users", [r.users for r in schedule], ""))

        # Uplift
        metric_rows.append(("Uplift", [r.uplift_pct * 100 for r in schedule], ""))

        # Product rates (recurring)
        for name in schedule[0].product_rates:
            metric_rows.append((f"{name} (Rate)", [r.product_rates.get(name, 0) for r in schedule], ""))

        # ARR
        arr_vals = [r.arr for r in schedule]
        metric_rows.append(("ARR", arr_vals, sum(arr_vals)))

        # One-time charges
        for key in schedule[0].one_time_charges:
            ot_vals = [r.one_time_charges.get(key, 0) for r in schedule]
            metric_rows.append((key, ot_vals, sum(ot_vals)))

        # Total Billed
        billed_vals = [r.total_billed for r in schedule]
        metric_rows.append(("Total Billed", billed_vals, sum(billed_vals)))

        dollar_labels = {"ARR", "Total Billed"}
        dollar_labels.update(schedule[0].one_time_charges.keys())
        dollar_labels.update(f"{n} (Rate)" for n in schedule[0].product_rates)

        for label, vals, total_val in metric_rows:
            is_total_row = label in ("ARR", "Total Billed")
            cell = ws.cell(row=row, column=1, value=label)
            if is_total_row:
                cell.fill = TOTAL_FILL
                cell.font = TOTAL_FONT
            for c, v in enumerate(vals, 2):
                cell = ws.cell(row=row, column=c, value=v)
                if label in dollar_labels:
                    cell.number_format = MONEY_FMT
                elif label == "Uplift":
                    cell.number_format = UPLIFT_FMT
                elif label == "Users":
                    cell.number_format = '#,##0'
                if is_total_row:
                    cell.fill = TOTAL_FILL
                    cell.font = TOTAL_FONT
            # Total column
            total_col = len(vals) + 2
            if total_val != "":
                cell = ws.cell(row=row, column=total_col, value=total_val)
                if label in dollar_labels:
                    cell.number_format = MONEY_FMT
                if is_total_row:
                    cell.fill = TOTAL_FILL
                    cell.font = TOTAL_FONT
            row += 1

        row += 3  # spacer

    # ── Billed Savings table ──
    if len(export_scenarios) >= 2:
        baseline = export_scenarios[0]
        target   = export_scenarios[1]
        shared   = min(baseline["Term"], target["Term"])

        billed_rows: list[dict] = []
        total_base = total_tgt = 0.0
        for yr in range(shared):
            base_yr = baseline["FullSchedule"][yr]
            tgt_yr  = target["FullSchedule"][yr]
            saving  = base_yr.total_billed - tgt_yr.total_billed
            total_base += base_yr.total_billed
            total_tgt  += tgt_yr.total_billed
            billed_rows.append({
                "Year": f"Year {yr + 1}",
                baseline["Label"]: base_yr.total_billed,
                target["Label"]:   tgt_yr.total_billed,
                "Saving ($)": saving,
                "Saving (%)": saving / base_yr.total_billed if base_yr.total_billed else 0,
            })
        total_saving = total_base - total_tgt
        billed_rows.append({
            "Year": "Total",
            baseline["Label"]: total_base,
            target["Label"]:   total_tgt,
            "Saving ($)": total_saving,
            "Saving (%)": total_saving / total_base if total_base else 0,
        })

        billed_cols = list(billed_rows[0].keys())
        write_header(billed_cols)
        for br in billed_rows:
            is_total = br["Year"] == "Total"
            for c, col_name in enumerate(billed_cols, 1):
                cell = ws.cell(row=row, column=c, value=br[col_name])
                apply_fmt(cell, col_name, is_total=is_total)
            row += 1

    # ── Auto-fit column widths ──
    for col in ws.columns:
        width = max((len(str(cell.value)) if cell.value is not None else 0) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 4, 32)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# Streamlit UI
# ══════════════════════════════════════════════════════════════════════════════

CSS = """
<style>
/* ═══ Deal Desk — Apple/Slack-inspired design ═══
   Font: system stack (SF Pro on Mac, Segoe on Windows)
   Scale: 11px labels · 13px body · 15px subhead · 20px title
   Palette: #1D1D1F text · #86868B secondary · #007AFF accent
   Spacing: 8px grid
*/

/* ── Typography ── */
.stMarkdown, .stMarkdown p, .stMarkdown h1, .stMarkdown h2, .stMarkdown h3,
.stMarkdown h4, .stMarkdown h5, .stMarkdown li, label,
th, td, input, select, textarea,
.stButton > button, [data-testid="stBaseButton-secondary"] {
    font-family: -apple-system, BlinkMacSystemFont, 'SF Pro Text',
                 'Segoe UI', Roboto, Helvetica, Arial, sans-serif !important;
    -webkit-font-smoothing: antialiased !important;
}

/* ── Kill SVG icon text leaks (arrow_right etc.) ── */
svg title, svg desc, svg text,
[data-testid="stExpander"] svg *,
[data-testid="stExpanderToggleIcon"] * {
    font-size: 0 !important; color: transparent !important;
    overflow: hidden !important; line-height: 0 !important;
}
[data-testid="stExpanderToggleIcon"] {
    overflow: hidden !important; width: 20px !important; height: 20px !important;
    flex-shrink: 0 !important;
}

/* ── Page background ── */
.stApp, .block-container, header, [data-testid="stHeader"] { background: #FFFFFF !important; }
.block-container { padding-top: 1rem !important; max-width: 1400px !important; }

/* ── Ensure select dropdowns never show a text cursor ── */
[data-baseweb="select"] input { caret-color: transparent !important; cursor: pointer !important; }
#MainMenu, footer, [data-testid="stDecoration"] { display: none !important; }

/* ── Brand ── */
.brand { display: flex; align-items: center; gap: 10px; margin-bottom: 8px; overflow: visible !important; }
.brand-logo { font-size: 22px; font-weight: 700; letter-spacing: -0.3px; color: #1D1D1F !important; white-space: nowrap !important; }
.brand-divider { width: 1px; height: 20px; background: #D2D2D7; }
.brand-sub { font-size: 14px !important; font-weight: 400 !important; color: #86868B !important; }

/* ── Metrics ── */
[data-testid="stMetric"] {
    background: #FFFFFF !important;
    border: 1px solid #E5E5EA !important; border-left: 3px solid #007AFF !important;
    padding: 16px 20px !important; border-radius: 12px !important;
    box-shadow: 0 1px 2px rgba(0,0,0,0.04) !important;
}
[data-testid="stMetricValue"] div { font-size: 24px !important; font-weight: 600 !important; color: #1D1D1F !important; }
[data-testid="stMetricLabel"] div { font-size: 11px !important; font-weight: 600 !important; color: #86868B !important; text-transform: uppercase !important; letter-spacing: 0.6px !important; }

/* ═══ SIDEBAR ═══ */
[data-testid="stSidebar"] { background: #F5F5F7 !important; border-right: 1px solid #E5E5EA !important; }
[data-testid="stSidebar"] [data-testid="stSidebarContent"] { padding: 8px 16px !important; }
[data-testid="stSidebar"] [data-testid="stSidebarUserContent"] { padding-top: 0 !important; }

/* Section headers */
[data-testid="stSidebar"] .stMarkdown h3 {
    font-size: 11px !important; font-weight: 600 !important; text-transform: uppercase !important;
    letter-spacing: 0.8px !important; color: #86868B !important;
    margin: 16px 0 8px 0 !important; padding: 0 !important;
}

/* Dividers — very subtle */
[data-testid="stSidebar"] hr { border-color: #E5E5EA !important; margin: 16px 0 !important; opacity: 0.5; }

/* Labels */
[data-testid="stSidebar"] label {
    color: #1D1D1F !important; font-size: 13px !important; font-weight: 500 !important;
    overflow: hidden !important; text-overflow: ellipsis !important;
    white-space: nowrap !important;
}
[data-testid="stSidebar"] .stMarkdown p {
    font-size: 13px !important; color: #1D1D1F !important;
    overflow-wrap: break-word !important; word-wrap: break-word !important;
}

/* Text inputs (exclude inputs inside selects — those cause caret bleed) */
[data-testid="stSidebar"] [data-baseweb="input"] input,
[data-testid="stSidebar"] [data-testid="stNumberInput"] input {
    background: #FFFFFF !important; border: 1px solid #D2D2D7 !important;
    color: #1D1D1F !important; border-radius: 8px !important; font-size: 14px !important;
}
[data-testid="stSidebar"] [data-baseweb="input"] input:focus,
[data-testid="stSidebar"] [data-testid="stNumberInput"] input:focus {
    border-color: #007AFF !important;
    box-shadow: 0 0 0 3px rgba(0,122,255,0.1) !important;
}

/* Selects / dropdowns */
[data-testid="stSidebar"] [data-baseweb="select"] > div {
    background: #FFFFFF !important; border-color: #D2D2D7 !important;
    border-radius: 8px !important;
}

/* Number steppers */
[data-testid="stSidebar"] [data-testid="stNumberInputStepUp"],
[data-testid="stSidebar"] [data-testid="stNumberInputStepDown"] {
    color: #86868B !important; border-color: #D2D2D7 !important;
}

/* Radio buttons */
.stRadio > div { gap: 8px !important; }

/* Fee/product cards */
[data-testid="stSidebar"] [data-testid="stVerticalBlockBorderWrapper"] {
    background: #FFFFFF !important; border: 1px solid #E5E5EA !important;
    border-radius: 10px !important; margin-bottom: 8px !important;
}

/* Sidebar primary buttons (Add Fee, Add Product, Clear) */
[data-testid="stSidebar"] .stButton > button {
    background: #FFFFFF !important; color: #007AFF !important;
    border: 1px solid #D2D2D7 !important; border-radius: 8px !important;
    font-size: 13px !important; padding: 6px 16px !important; font-weight: 500 !important;
}
[data-testid="stSidebar"] .stButton > button:hover {
    background: #F5F5F7 !important; transform: none !important;
}
[data-testid="stSidebar"] .stButton > button * { color: #007AFF !important; }

/* Sidebar delete (✕) buttons */
[data-testid="stSidebar"] [data-testid="stBaseButton-secondary"] {
    background: transparent !important; border: 1px solid #D2D2D7 !important;
    color: #AEAEB2 !important; padding: 2px 8px !important; min-height: 0 !important;
    border-radius: 6px !important;
}
[data-testid="stSidebar"] [data-testid="stBaseButton-secondary"]:hover {
    background: #FFF0F0 !important; border-color: #FFD4D4 !important;
}
[data-testid="stSidebar"] [data-testid="stBaseButton-secondary"] * { color: #AEAEB2 !important; }
[data-testid="stSidebar"] [data-testid="stBaseButton-secondary"]:hover * { color: #FF3B30 !important; }

/* Sidebar slider */
[data-testid="stSidebar"] [data-baseweb="slider"] [role="slider"] { background: #007AFF !important; }

/* Sidebar spacing */
[data-testid="stSidebar"] [data-testid="stVerticalBlock"] { gap: 0.5rem !important; }

/* ═══ MAIN AREA ═══ */

/* Primary buttons */
.stButton > button {
    background: #007AFF !important; color: #FFF !important;
    border: none !important; font-weight: 500 !important; border-radius: 8px !important;
    font-size: 14px !important; padding: 8px 20px !important;
}
.stButton > button:hover { background: #0066CC !important; transform: none !important; }
.stButton > button * { color: #FFF !important; }

/* Secondary buttons (✕ in main area) */
[data-testid="stBaseButton-secondary"] {
    background: transparent !important; color: #AEAEB2 !important;
    border: 1px solid #E5E5EA !important; border-radius: 6px !important;
    padding: 2px 10px !important; font-size: 13px !important;
    min-height: 0 !important; box-shadow: none !important;
}
[data-testid="stBaseButton-secondary"]:hover {
    background: #FFF0F0 !important; border-color: #FFD4D4 !important;
}
[data-testid="stBaseButton-secondary"] * { color: #AEAEB2 !important; }
[data-testid="stBaseButton-secondary"]:hover * { color: #FF3B30 !important; }

/* Main area text inputs & selects */
[data-baseweb="select"] > div { border-radius: 8px !important; }
[data-baseweb="input"] { border-radius: 8px !important; }

/* Expander */
[data-testid="stExpander"] {
    border: 1px solid #E5E5EA !important; border-radius: 10px !important;
    overflow: hidden !important;
}
[data-testid="stExpander"] summary,
[data-testid="stExpander"] [data-testid="stExpanderDetails"] > div > div > p {
    font-size: 13px !important; font-weight: 500 !important;
    color: #86868B !important; overflow: hidden !important;
    text-overflow: ellipsis !important;
}
[data-testid="stExpander"] summary {
    background: transparent !important;
}
[data-testid="stExpander"] svg {
    overflow: hidden !important; width: 16px !important; height: 16px !important;
}

/* ═══ DATA TABLES ═══ */
[data-testid="stDataFrame"] { border-radius: 10px !important; overflow: hidden !important; }
th {
    background: #F5F5F7 !important; font-weight: 500 !important; font-size: 11px !important;
    text-transform: uppercase !important; letter-spacing: 0.5px !important;
    color: #86868B !important; border-bottom: 1px solid #E5E5EA !important;
}

/* ═══ TABS ═══ */
.stTabs [data-baseweb="tab-list"] { gap: 0; border-bottom: 1px solid #E5E5EA; }
.stTabs [data-baseweb="tab"] {
    font-weight: 500 !important; font-size: 14px !important;
    padding: 12px 24px !important; color: #86868B !important;
    border-bottom: 2px solid transparent; margin-bottom: -1px;
}
.stTabs [aria-selected="true"] {
    color: #1D1D1F !important; border-bottom-color: #007AFF !important;
    font-weight: 600 !important; background: transparent !important;
}

/* ═══ CUSTOM CLASSES ═══ */
.scenario-header { font-size: 11px; font-weight: 600; text-transform: uppercase; letter-spacing: 0.8px; color: #86868B; margin-bottom: 8px; }
.section-header { font-size: 20px !important; font-weight: 600 !important; color: #1D1D1F !important; margin-bottom: 4px !important; letter-spacing: -0.2px !important; }
.section-sub { font-size: 14px !important; color: #86868B !important; margin-bottom: 24px !important; }
.stat-row { display: flex; gap: 32px; margin-top: 16px; }
.stat-item { display: flex; flex-direction: column; }
.stat-label { font-size: 11px; font-weight: 600; text-transform: uppercase; letter-spacing: 0.5px; color: #86868B; }
.stat-val { font-size: 16px; font-weight: 600; color: #1D1D1F; margin-top: 2px; }
[data-testid="stAlert"] { border-radius: 10px !important; font-size: 14px !important; }
</style>
"""

# ── Page Config ──────────────────────────────────────────────────────────────

st.set_page_config(page_title="Deal Desk Modeler", layout="wide")
st.markdown(CSS, unsafe_allow_html=True)
st.markdown(
    '<div class="brand">'
    '<span class="brand-logo">Deal Desk</span>'
    '<span class="brand-divider"></span>'
    '<span class="brand-sub">Pricing Modeler</span>'
    "</div>",
    unsafe_allow_html=True,
)


# ── Session State ────────────────────────────────────────────────────────────


def _init_state() -> None:
    if "portfolio" not in st.session_state:
        st.session_state.portfolio = [
            {"name": p.name, "type": p.pricing_type.value, "price": p.base_price}
            for p in DEFAULT_PRODUCTS
        ]
    if "one_time_fees" not in st.session_state:
        st.session_state.one_time_fees = []
    for idx, cfg in enumerate(SCENARIO_DEFAULTS):
        defaults = {
            f"s_term_{idx}": cfg.term,
            f"s_name_{idx}": cfg.name,
            f"s_pay_{idx}": cfg.payment_label,
            f"s_mode_{idx}": cfg.mode.value,
            f"s_add_{idx}": list(cfg.added_fees),
            f"s_ledger_{idx}": empty_discount_ledger(),
        }
        for key, value in defaults.items():
            if key not in st.session_state:
                st.session_state[key] = value


_init_state()


# ── Converters ───────────────────────────────────────────────────────────────


def _products_from_state() -> list[Product]:
    """Recurring products contribute to ARR every year."""
    return [
        Product(p["name"], PricingType(p["type"]), p["price"])
        for p in st.session_state.portfolio
        if p.get("cadence", "One-Time") == "Recurring"
    ]


def _fees_from_state() -> list[Fee]:
    """One-time products are treated as fees (Year 1 only)."""
    return [
        Fee(p["name"], p["price"])
        for p in st.session_state.portfolio
        if p.get("cadence", "One-Time") == "One-Time"
    ]


# ── Sidebar ──────────────────────────────────────────────────────────────────

with st.sidebar:
    st.markdown("### Deal Setup")
    num_scenarios = st.selectbox("Scenarios", options=[1, 2, 3, 4, 5], index=1)

    st.markdown("---")
    st.markdown("### Users")
    growth = st.radio(
        "Model", ["Flat", "Custom Ramp"], horizontal=True, label_visibility="collapsed"
    )

    if growth == "Flat":
        flat_users = st.number_input(
            "Users (all years)", value=DEFAULT_USERS, step=100, min_value=0
        )
        user_ramp = [flat_users] * MAX_TERM
    else:
        ramp_years = st.slider("Years", 1, MAX_TERM, 3, key="ramp_yrs")
        ramp_values: list[int] = []
        col_a, col_b = st.columns(2)
        for yr in range(ramp_years):
            with col_a if yr % 2 == 0 else col_b:
                ramp_values.append(
                    st.number_input(
                        f"Yr {yr + 1}", value=DEFAULT_USERS, step=100, min_value=0, key=f"ramp_{yr}"
                    )
                )
        user_ramp = pad_to_term(
            [float(v) for v in ramp_values], MAX_TERM, fill=float(DEFAULT_USERS)
        )
        user_ramp = [int(u) for u in user_ramp]

    # ── Products ──
    st.markdown("---")
    st.markdown("### Products")

    # Search-and-add: selectbox with available (not yet added) products
    added_names = {p["name"] for p in st.session_state.portfolio}
    available = [""] + [n for n in CATALOG_NAMES if n not in added_names]

    pick = st.selectbox(
        "Add a product",
        options=available,
        index=0,
        key="product_search",
        format_func=lambda x: "Type to search..." if x == "" else x,
    )
    if pick:
        cat_entry = CATALOG_LOOKUP.get(pick, {})
        st.session_state.portfolio.append({
            "name": pick,
            "type": PricingType.FLAT_FEE.value,
            "price": cat_entry.get("list_price", 0.0),
            "list_price": cat_entry.get("list_price", 0.0),
            "cadence": cat_entry.get("cadence", "One-Time"),
        })
        st.rerun()

    # Product cards
    delete_prod_idx: int | None = None
    for idx, prod in enumerate(st.session_state.portfolio):
        if "list_price" not in prod:
            cat_entry = CATALOG_LOOKUP.get(prod["name"], {})
            prod["list_price"] = cat_entry.get("list_price", 0.0)
            if "cadence" not in prod:
                prod["cadence"] = cat_entry.get("cadence", "One-Time")

        with st.container(border=True):
            # Header: bold name + cadence badge
            cadence = prod.get("cadence", "One-Time")
            cad_color = "#007AFF" if cadence == "Recurring" else "#86868B"
            st.markdown(
                f'<span style="font-size:13px;color:#1D1D1F;font-weight:700;">{prod["name"]}</span>'
                f' <span style="font-size:10px;color:{cad_color};font-weight:600;'
                f'text-transform:uppercase;letter-spacing:0.5px;">{cadence}</span>',
                unsafe_allow_html=True,
            )

            type_col, price_col, del_col = st.columns([3, 3, 1])
            with type_col:
                type_options = [t.value for t in PricingType]
                prod["type"] = st.selectbox(
                    "Type", type_options, index=type_options.index(prod["type"]), key=f"pt_{idx}"
                )
            with price_col:
                prod["price"] = st.number_input(
                    "Price ($)", value=prod["price"], key=f"pp_{idx}", format="%.2f", min_value=0.0
                )
            with del_col:
                st.markdown("<div style='height:27px'></div>", unsafe_allow_html=True)
                if st.button("✕", key=f"pd_{idx}", type="secondary"):
                    delete_prod_idx = idx

            # Variance from list price
            lp = prod["list_price"]
            cp = prod["price"]
            if lp > 0 and cp != lp:
                diff_pct = ((cp - lp) / lp) * 100
                if cp < lp:
                    st.markdown(
                        f'<span style="color:#FF3B30;font-size:12px;">'
                        f'▼ {abs(diff_pct):.1f}% below list (${lp:,.2f})</span>',
                        unsafe_allow_html=True,
                    )
                else:
                    st.markdown(
                        f'<span style="color:#34C759;font-size:12px;">'
                        f'▲ {diff_pct:.1f}% above list (${lp:,.2f})</span>',
                        unsafe_allow_html=True,
                    )
            elif lp > 0:
                st.markdown(
                    f'<span style="color:#86868B;font-size:12px;">At list price</span>',
                    unsafe_allow_html=True,
                )

    if delete_prod_idx is not None:
        st.session_state.portfolio.pop(delete_prod_idx)
        st.rerun()


# ── Scenarios ────────────────────────────────────────────────────────────────

product_names = [p["name"] for p in st.session_state.portfolio]
products = _products_from_state()
fees = _fees_from_state()
available_fee_names = [f.name for f in fees]

results: list[dict] = []
scenario_cols = st.columns(num_scenarios, gap="large")

for scene_idx, col in enumerate(scenario_cols):
    with col:
        st.markdown(
            f'<p class="scenario-header">Scenario {scene_idx + 1}</p>',
            unsafe_allow_html=True,
        )
        label = st.text_input(
            "Label", value=st.session_state[f"s_name_{scene_idx}"],
            key=f"wn_{scene_idx}", label_visibility="collapsed", placeholder="Scenario name…",
        )
        st.session_state[f"s_name_{scene_idx}"] = label

        term_col, pay_col, mode_col = st.columns(3)
        with term_col:
            term_options = list(range(1, MAX_TERM + 1))
            term = st.selectbox(
                "Term", term_options,
                index=term_options.index(st.session_state[f"s_term_{scene_idx}"]),
                key=f"wt_{scene_idx}",
            )
            st.session_state[f"s_term_{scene_idx}"] = term

        with pay_col:
            pay_options = list(PAY_TERMS.keys())
            payment = st.selectbox(
                "Payment", pay_options,
                index=pay_options.index(st.session_state[f"s_pay_{scene_idx}"]),
                key=f"wp_{scene_idx}",
            )
            st.session_state[f"s_pay_{scene_idx}"] = payment

        with mode_col:
            mode_options = [m.value for m in UpliftMode]
            current_mode = st.session_state[f"s_mode_{scene_idx}"]
            mode_index = mode_options.index(current_mode) if current_mode in mode_options else 0
            mode = st.selectbox("Mode", mode_options, index=mode_index, key=f"wm_{scene_idx}")
            st.session_state[f"s_mode_{scene_idx}"] = mode

        # All one-time products auto-included in Year 1
        added = available_fee_names

        payment_days = PAY_TERMS[payment]
        ledger = st.session_state[f"s_ledger_{scene_idx}"]

        with st.expander("Custom Line-Item Discounts", expanded=False):
            ledger_config = {
                "Year": st.column_config.NumberColumn("Year", min_value=1, max_value=MAX_TERM, step=1, required=True),
                "Product": st.column_config.SelectboxColumn("Product", options=product_names, required=True),
                "Discount (%)": st.column_config.NumberColumn("%", min_value=0.0, max_value=100.0, step=0.1, required=True),
            }
            ledger = st.data_editor(
                ledger, column_config=ledger_config, num_rows="dynamic",
                use_container_width=True, key=f"ed_{scene_idx}", hide_index=True,
            )
            st.session_state[f"s_ledger_{scene_idx}"] = ledger

        # ── Uplift Config ──
        uplifts: list[float] = []

        if mode == UpliftMode.GOAL_SEEK.value:
            if f"s_target_{scene_idx}" not in st.session_state:
                st.session_state[f"s_target_{scene_idx}"] = 250_000.0
            target = st.number_input(
                "Target TCV ($)", value=st.session_state[f"s_target_{scene_idx}"],
                step=1000.0, key=f"wtv_{scene_idx}", min_value=0.0,
            )
            st.session_state[f"s_target_{scene_idx}"] = target

            if not products:
                st.warning("Add a product first.")
                solved = 0.0
            else:
                solved = goal_seek_uplift(
                    target_tcv=target, users=user_ramp, products=products, fees=fees,
                    added_fee_names=added, payment_days=payment_days,
                    ledger=ledger, term=term,
                )

            if solved < -0.20:
                st.warning(f"Steep discount needed: {solved * 100:.1f}% — verify target.")
            else:
                st.caption(f"→ Solved uplift: **{solved * 100:.2f}%**")
            uplifts = [solved] * term

        else:
            variable = st.toggle("Variable uplifts", value=False, key=f"vt_{scene_idx}")
            if not variable:
                flat_uplift = (
                    st.slider("Flat Uplift %", UPLIFT_MIN, UPLIFT_MAX, 0, key=f"fu_{scene_idx}") / 100
                )
                uplifts = [flat_uplift] * term
            else:
                for yr_idx in range(term):
                    uplifts.append(
                        st.slider(
                            f"Yr {yr_idx + 1} %", UPLIFT_MIN, UPLIFT_MAX, 0,
                            key=f"vu_{scene_idx}_{yr_idx}",
                        ) / 100
                    )

        # ── Run Model ──
        full_schedule = run_model(
            users=user_ramp, products=products, fees=fees, added_fee_names=added,
            uplifts=uplifts, payment_days=payment_days, ledger=ledger,
        )
        schedule = full_schedule[:term]

        tcv = sum(r.total_billed for r in schedule)
        ending_arr = schedule[-1].arr if schedule else 0

        st.metric("Nominal TCV", f"${tcv:,.0f}")
        st.markdown(
            f'<div class="stat-row">'
            f'<div class="stat-item"><span class="stat-label">Ending ARR</span>'
            f'<span class="stat-val">${ending_arr:,.0f}</span></div>'
            f"</div>",
            unsafe_allow_html=True,
        )

        results.append({
            "Label": label,
            "Schedule": schedule,
            "FullSchedule": full_schedule,
            "TCV": tcv,
            "Ending ARR": ending_arr,
            "Term": term,
            "Added Fees": added,
        })


# ── Detail Tabs ──────────────────────────────────────────────────────────────

st.markdown("")
st.markdown("")

tab_names = ["📊 Breakdown"]
if num_scenarios > 1:
    tab_names.append(f"⚖️ vs. {results[0]['Label']}")
tab_names.append("🏢 EE Pricing Calculator")
tab_names.append("💲 Pricing Table")
tab_names.append("📥 Export")
tabs = st.tabs(tab_names)

# ── Tab: Breakdown ──
with tabs[0]:
    st.markdown('<p class="section-header">Financial Breakdown</p>', unsafe_allow_html=True)
    st.markdown('<p class="section-sub">Year-by-year schedule per scenario</p>', unsafe_allow_html=True)
    for idx in range(num_scenarios):
        st.markdown(f"**{results[idx]['Label']}**")
        df = results_to_dataframe(results[idx]["Schedule"])

        # Build format dict: currency for dollar rows, comma for user rows
        dollar_labels = {"ARR", "Total Billed"}
        # Add one-time charge keys and rate keys
        for r in results[idx]["Schedule"]:
            dollar_labels.update(r.one_time_charges.keys())
            dollar_labels.update(f"{n} (Rate)" for n in r.product_rates)

        def fmt_cell(val, row_label):
            if isinstance(val, str) or val == "" or val is None:
                return val
            if row_label == "Users":
                return f"{int(val):,}"
            if row_label in dollar_labels:
                return f"${val:,.2f}"
            return val

        styled_data = []
        for _, row in df.iterrows():
            label = row[""]
            styled_row = {"": label}
            for col in df.columns[1:]:
                styled_row[col] = fmt_cell(row[col], label)
            styled_data.append(styled_row)

        display_df = pd.DataFrame(styled_data)

        def highlight_totals(row: pd.Series) -> list[str]:
            bold = "font-weight:600;background:#F1F5F9"
            if row[""] in ("ARR", "Total Billed"):
                return [bold] * len(row)
            return [""] * len(row)

        st.dataframe(
            display_df.style.apply(highlight_totals, axis=1),
            use_container_width=True,
            hide_index=True,
        )
        if idx < num_scenarios - 1:
            st.divider()

# ── Tab: Comparison ──
if num_scenarios > 1:
    with tabs[1]:
        st.markdown('<p class="section-header">Deal Comparison</p>', unsafe_allow_html=True)
        st.markdown(
            '<p class="section-sub">Each scenario measured against the baseline</p>',
            unsafe_allow_html=True,
        )

        baseline = results[0]
        for comp_idx in range(1, num_scenarios):
            target_scenario = results[comp_idx]
            shared_years = min(baseline["Term"], target_scenario["Term"])
            st.markdown(
                f"**{target_scenario['Label']}** vs. **{baseline['Label']}** "
                f"— first {shared_years} yr{'s' if shared_years > 1 else ''}"
            )

            per_user_rows: list[dict] = []
            billed_rows: list[dict] = []
            total_baseline_billed = 0.0
            total_target_billed = 0.0

            for yr in range(shared_years):
                base_yr = baseline["FullSchedule"][yr]
                tgt_yr = target_scenario["FullSchedule"][yr]
                user_count = tgt_yr.users

                base_per_user = base_yr.total_billed / user_count if user_count else 0
                tgt_per_user = tgt_yr.total_billed / user_count if user_count else 0
                saving_per_user = base_per_user - tgt_per_user

                per_user_rows.append({
                    "Year": f"Year {yr + 1}",
                    "Blended $/User": tgt_per_user,
                    f"Savings vs. {baseline['Label']}": saving_per_user,
                    "Saving %": saving_per_user / base_per_user if base_per_user else 0,
                })

                base_billed = base_yr.total_billed
                tgt_billed = tgt_yr.total_billed
                total_baseline_billed += base_billed
                total_target_billed += tgt_billed
                saving = base_billed - tgt_billed

                billed_rows.append({
                    "Year": f"Year {yr + 1}",
                    f"{baseline['Label']}": base_billed,
                    f"{target_scenario['Label']}": tgt_billed,
                    "Saving ($)": saving,
                    "Saving (%)": saving / base_billed if base_billed else 0,
                })

            total_saving = total_baseline_billed - total_target_billed
            billed_rows.append({
                "Year": "Total",
                f"{baseline['Label']}": total_baseline_billed,
                f"{target_scenario['Label']}": total_target_billed,
                "Saving ($)": total_saving,
                "Saving (%)": total_saving / total_baseline_billed if total_baseline_billed else 0,
            })

            st.caption("Per-User Blended Rate")
            st.dataframe(
                pd.DataFrame(per_user_rows).style.format({
                    "Blended $/User": "${:,.2f}",
                    f"Savings vs. {baseline['Label']}": "${:,.2f}",
                    "Saving %": "{:.2%}",
                }),
                use_container_width=True, hide_index=True,
            )

            st.caption("Billed Savings (Nominal)")
            billed_df = pd.DataFrame(billed_rows)

            def highlight_total(row: pd.Series) -> list[str]:
                style = "background:#F1F5F9;font-weight:600"
                return [style] * len(row) if row["Year"] == "Total" else [""] * len(row)

            st.dataframe(
                billed_df.style.apply(highlight_total, axis=1).format({
                    f"{baseline['Label']}": "${:,.2f}",
                    f"{target_scenario['Label']}": "${:,.2f}",
                    "Saving ($)": "${:,.2f}",
                    "Saving (%)": "{:.2%}",
                }),
                use_container_width=True, hide_index=True,
            )

            if comp_idx < num_scenarios - 1:
                st.markdown("---")

# ── Tab: EE Pricing Calculator ──
with tabs[-3]:
    st.markdown('<p class="section-header">EE Pricing Calculator</p>', unsafe_allow_html=True)
    st.markdown(
        '<p class="section-sub">Quick-reference pricing for Employee Engagement modules</p>',
        unsafe_allow_html=True,
    )

    # ── Session state defaults ──
    if "ee_users" not in st.session_state:
        st.session_state.ee_users = 0
    if "ee_region" not in st.session_state:
        st.session_state.ee_region = "USA"
    if "ee_impl_override" not in st.session_state:
        st.session_state.ee_impl_override = 0.0

    # Module selections: Essential / Advanced / None
    if "ee_platform" not in st.session_state:
        st.session_state.ee_platform = "Essential"
    if "ee_donate" not in st.session_state:
        st.session_state.ee_donate = "Essential"
    if "ee_volunteer" not in st.session_state:
        st.session_state.ee_volunteer = "—"
    if "ee_eg" not in st.session_state:
        st.session_state.ee_eg = "—"

    # ── Reference data (editable in expander) ──
    if "ee_regions" not in st.session_state:
        st.session_state.ee_regions = {
            "USA":    {"adj": 1.0},
            "US Int'l": {"adj": 1.0},
            "CAN":    {"adj": 1.0},
            "UK":     {"adj": 0.7395},
            "EUR (Rest)": {"adj": 0.75375},
            "FRA":    {"adj": 0.7},
            "GER":    {"adj": 0.804},
            "CHF":    {"adj": 0.820845},
            "ANZ":    {"adj": 1.100328},
            "NBNLUX": {"adj": 0.75375},
        }

    if "ee_tiers" not in st.session_state:
        # cumul = cumulative total at the START of this tier (i.e. total for all prior tiers)
        st.session_state.ee_tiers = [
            {"low": 1,       "high": 2158,     "rate": 18.528058, "cumul": 0.0},
            {"low": 2159,    "high": 2999,     "rate": 13.18,     "cumul": 40002.077},
            {"low": 3000,    "high": 4999,     "rate": 12.16,     "cumul": 51086.470},
            {"low": 5000,    "high": 9999,     "rate": 5.68,      "cumul": 75406.459},
            {"low": 10000,   "high": 24999,    "rate": 4.26,      "cumul": 103806.459},
            {"low": 25000,   "high": 49999,    "rate": 2.83,      "cumul": 167706.459},
            {"low": 50000,   "high": 74999,    "rate": 2.09,      "cumul": 238456.459},
            {"low": 75000,   "high": 99999,    "rate": 2.09,      "cumul": 290706.459},
            {"low": 100000,  "high": 149999,   "rate": 2.09,      "cumul": 342956.459},
            {"low": 150000,  "high": 249999,   "rate": 1.33,      "cumul": 447456.459},
            {"low": 250000,  "high": 499999,   "rate": 0.91,      "cumul": 580456.459},
            {"low": 500000,  "high": 1500000,  "rate": 0.91,      "cumul": 807956.459},
        ]

    if "ee_module_weights" not in st.session_state:
        st.session_state.ee_module_weights = {
            "Platform Essential": 0.39, "Platform Advanced": 0.50,
            "Donate Essential":   0.20, "Donate Advanced":   0.25,
            "Volunteer Essential": 0.25, "Volunteer Advanced": 0.50,
            "EG Essential":       0.18, "EG Advanced":       0.36,
        }

    if "ee_min_floors" not in st.session_state:
        # Floors per module per tier band: [up to 5K, 5-10K, 10-50K]
        st.session_state.ee_min_floors = {
            "Platform Essential": [15600.81, 19923.72, 29408.52],
            "Platform Advanced":  [20001.04, 25543.24, 37703.23],
            "Donate Essential":   [8000.42,  10217.29, 15081.29],
            "Donate Advanced":    [10000.52, 12771.62, 18851.61],
            "Volunteer Essential": [10000.52, 12771.62, 18851.61],
            "Volunteer Advanced":  [20001.04, 25543.24, 37703.23],
            "EG Essential":       [7200.37,  9195.56,  13573.16],
            "EG Advanced":        [14400.75, 18391.13, 27146.33],
        }

    if "ee_impl_costs" not in st.session_state:
        # [up to 5K, 5-10K, 10-50K, 50K+]
        st.session_state.ee_impl_costs = {
            "Platform Essential": [5500,  6667,  10000, 15000],
            "Platform Advanced":  [8250,  10000, 15000, 22500],
            "Donate Essential":   [6875,  8333,  12500, 18750],
            "Donate Advanced":    [11000, 13333, 20000, 30000],
            "Volunteer Essential": [1375,  1667,  2500,  3750],
            "Volunteer Advanced":  [2750,  3333,  5000,  7500],
            "EG Essential":       [3300,  4000,  6000,  9000],
            "EG Advanced":        [8250,  10000, 15000, 22500],
        }

    # ── Helper: compute base annual amount from graduated tiers ──
    def ee_calc_base(users: int, region_adj: float) -> float:
        """Find the tier the user falls in, use cumulative + marginal rate."""
        if users <= 0:
            return 0.0
        tiers = st.session_state.ee_tiers
        # Find the tier where low <= users (last match)
        matched = tiers[0]
        for t in tiers:
            if t["low"] <= users:
                matched = t
            else:
                break
        total = matched["cumul"] + (users - matched["low"]) * matched["rate"]
        # If users exceed last tier, extend at last tier rate
        if users > tiers[-1]["high"]:
            total = tiers[-1]["cumul"] + (users - tiers[-1]["low"]) * tiers[-1]["rate"]
        return max(40000, total) * region_adj

    # ── Helper: get employee band index ──
    def ee_band_index(users: int) -> int:
        if users < 5000:
            return 0
        elif users < 10000:
            return 1
        elif users < 50000:
            return 2
        else:
            return 3

    def ee_band_label(users: int) -> str:
        labels = ["Up to 5K", "5-10K", "10-50K", "50K+"]
        return labels[ee_band_index(users)]

    # ── Inputs ──
    in1, in2 = st.columns(2)
    with in1:
        ee_users = st.number_input("Eligible Users", min_value=0, value=st.session_state.ee_users,
                                    step=100, key="ee_users_input")
        st.session_state.ee_users = ee_users
    with in2:
        region_names = list(st.session_state.ee_regions.keys())
        ee_region = st.selectbox("Region", region_names,
                                  index=region_names.index(st.session_state.ee_region),
                                  key="ee_region_input")
        st.session_state.ee_region = ee_region

    st.markdown("---")

    # ── Module tier selection ──
    st.markdown("**Select module tiers**")
    mod_options = ["—", "Essential", "Advanced"]
    mc1, mc2, mc3, mc4 = st.columns(4)
    with mc1:
        ee_platform = st.selectbox("Platform", mod_options,
                                    index=mod_options.index(st.session_state.ee_platform),
                                    key="ee_platform_sel")
        st.session_state.ee_platform = ee_platform
    with mc2:
        donate_opts = ["—", "Essential", "Advanced"]
        ee_donate = st.selectbox("Donate", donate_opts,
                                  index=donate_opts.index(st.session_state.ee_donate),
                                  key="ee_donate_sel")
        st.session_state.ee_donate = ee_donate
    with mc3:
        vol_opts = ["—", "Essential"]
        ee_volunteer = st.selectbox("Volunteer", vol_opts,
                                     index=vol_opts.index(st.session_state.ee_volunteer) if st.session_state.ee_volunteer in vol_opts else 0,
                                     key="ee_volunteer_sel")
        st.session_state.ee_volunteer = ee_volunteer
    with mc4:
        ee_eg = st.selectbox("Employee Groups", mod_options,
                              index=mod_options.index(st.session_state.ee_eg),
                              key="ee_eg_sel")
        st.session_state.ee_eg = ee_eg

    st.markdown("---")

    # ── Compute pricing ──
    region_adj = st.session_state.ee_regions[ee_region]["adj"]
    base_amount = ee_calc_base(ee_users, region_adj)
    band_idx = ee_band_index(ee_users)

    weights = st.session_state.ee_module_weights
    floors = st.session_state.ee_min_floors

    modules_selected = {
        "Platform": ee_platform,
        "Donate": ee_donate,
        "Volunteer": ee_volunteer,
        "EG": ee_eg,
    }

    module_rows = []
    total_annual = 0.0
    total_impl = 0.0

    for mod_name, tier_choice in modules_selected.items():
        if tier_choice == "—":
            continue
        key = f"{mod_name} {tier_choice}"
        weight = weights.get(key, 0)
        annual_sub = weight * base_amount

        # Apply minimum floor (floors only have 3 bands; 50K+ has no floor — use 10-50K)
        floor_idx = min(band_idx, 2)
        floor_val = floors.get(key, [0, 0, 0])[floor_idx]
        annual_sub = max(annual_sub, floor_val)

        user_rate = annual_sub / ee_users if ee_users else 0
        impl = st.session_state.ee_impl_costs.get(key, [0, 0, 0, 0])[band_idx]

        module_rows.append({
            "Module": mod_name,
            "Tier": tier_choice,
            "Annual Subscription": annual_sub,
            "User Rate": user_rate,
            "Implementation": impl,
        })
        total_annual += annual_sub
        total_impl += impl

    # ── Implementation override ──
    impl_col1, impl_col2 = st.columns(2)
    with impl_col1:
        impl_override = st.number_input(
            "Implementation Override ($)",
            min_value=0.0, value=st.session_state.ee_impl_override,
            step=1000.0, format="%.0f", key="ee_impl_override_input",
            help="Set to 0 to use standard implementation costs. Enter a custom value to override."
        )
        st.session_state.ee_impl_override = impl_override

    final_impl = impl_override if impl_override > 0 else total_impl

    # ── Summary metrics ──
    effective_rate = total_annual / ee_users if ee_users else 0
    year1_total = total_annual + final_impl

    m1, m2, m3 = st.columns(3)
    m1.metric("Annual Subscription", f"${total_annual:,.2f}")
    m2.metric("Effective $/User", f"${effective_rate:,.2f}")
    m3.metric("Year 1 Total", f"${year1_total:,.2f}")

    # ── Module breakdown table ──
    if module_rows:
        st.markdown("**Module Breakdown**")
        mod_df = pd.DataFrame(module_rows)
        # Add totals row
        totals_row = {
            "Module": "Total",
            "Tier": "",
            "Annual Subscription": total_annual,
            "User Rate": effective_rate,
            "Implementation": final_impl,
        }
        mod_df = pd.concat([mod_df, pd.DataFrame([totals_row])], ignore_index=True)

        def highlight_total_ee(row: pd.Series) -> list[str]:
            style = "background:#F1F5F9;font-weight:600"
            return [style] * len(row) if row["Module"] == "Total" else [""] * len(row)

        st.dataframe(
            mod_df.style.apply(highlight_total_ee, axis=1).format({
                "Annual Subscription": "${:,.2f}",
                "User Rate": "${:,.4f}",
                "Implementation": "${:,.0f}",
            }),
            use_container_width=True,
            hide_index=True,
        )

        # ── Year 1 total summary ──
        st.markdown("**Year 1 Summary**")
        y1c1, y1c2, y1c3 = st.columns(3)
        y1c1.metric("Implementation", f"${final_impl:,.0f}")
        y1c2.metric("Annual Subscription", f"${total_annual:,.2f}")
        y1c3.metric("Total Year 1", f"${year1_total:,.2f}")

    else:
        st.info("Select at least one module above to see pricing.")

    # ── Editable reference data (expander) ──
    with st.expander("Edit Reference Data", expanded=False):
        edit_tab1, edit_tab2, edit_tab3, edit_tab4 = st.tabs([
            "Graduated Tiers", "Module Weights", "Min Floors", "Implementation Costs"
        ])

        with edit_tab1:
            st.caption("Base per-user rates by tier (USD). Cumulative = total from all prior tiers.")
            for i, t in enumerate(st.session_state.ee_tiers):
                tc1, tc2, tc3, tc4 = st.columns([1, 1, 1, 1])
                with tc1:
                    new_low = st.number_input("Low", value=t["low"], key=f"eet_low_{i}",
                                               min_value=0, step=1)
                with tc2:
                    new_high = st.number_input("High", value=t["high"], key=f"eet_high_{i}",
                                                min_value=0, step=1)
                with tc3:
                    new_rate = st.number_input("Rate $/user", value=t["rate"], key=f"eet_rate_{i}",
                                                min_value=0.0, step=0.01, format="%.6f")
                with tc4:
                    new_cumul = st.number_input("Cumulative $", value=t["cumul"], key=f"eet_cumul_{i}",
                                                 min_value=0.0, step=100.0, format="%.3f")
                st.session_state.ee_tiers[i] = {"low": new_low, "high": new_high, "rate": new_rate, "cumul": new_cumul}

        with edit_tab2:
            st.caption("Each module's share of the base amount (decimal, e.g. 0.39 = 39%)")
            for wk, wv in st.session_state.ee_module_weights.items():
                new_w = st.number_input(wk, value=wv, min_value=0.0, max_value=1.0,
                                         step=0.01, format="%.4f", key=f"eew_{wk}")
                st.session_state.ee_module_weights[wk] = new_w

        with edit_tab3:
            st.caption("Minimum annual subscription floors by employee band [Up to 5K, 5-10K, 10-50K]")
            for fk, fv in st.session_state.ee_min_floors.items():
                st.markdown(f"**{fk}**")
                fc1, fc2, fc3 = st.columns(3)
                with fc1:
                    f0 = st.number_input("Up to 5K", value=fv[0], step=100.0, format="%.2f", key=f"eef_{fk}_0")
                with fc2:
                    f1 = st.number_input("5-10K", value=fv[1], step=100.0, format="%.2f", key=f"eef_{fk}_1")
                with fc3:
                    f2 = st.number_input("10-50K", value=fv[2], step=100.0, format="%.2f", key=f"eef_{fk}_2")
                st.session_state.ee_min_floors[fk] = [f0, f1, f2]

        with edit_tab4:
            st.caption("Implementation fees by employee band [Up to 5K, 5-10K, 10-50K, 50K+]")
            for ik, iv in st.session_state.ee_impl_costs.items():
                st.markdown(f"**{ik}**")
                ic1, ic2, ic3, ic4 = st.columns(4)
                with ic1:
                    i0 = st.number_input("Up to 5K", value=float(iv[0]), step=100.0, format="%.0f", key=f"eei_{ik}_0")
                with ic2:
                    i1 = st.number_input("5-10K", value=float(iv[1]), step=100.0, format="%.0f", key=f"eei_{ik}_1")
                with ic3:
                    i2 = st.number_input("10-50K", value=float(iv[2]), step=100.0, format="%.0f", key=f"eei_{ik}_2")
                with ic4:
                    i3 = st.number_input("50K+", value=float(iv[3]), step=100.0, format="%.0f", key=f"eei_{ik}_3")
                st.session_state.ee_impl_costs[ik] = [i0, i1, i2, i3]

    # ── Editable region adjustment factors ──
    with st.expander("Edit Region Adjustment Factors", expanded=False):
        st.caption("Adjustment factor applied to base amount (1.0 = no adjustment)")
        for rname, rdata in st.session_state.ee_regions.items():
            new_adj = st.number_input(rname, value=rdata["adj"], min_value=0.0,
                                       step=0.01, format="%.6f", key=f"eer_{rname}")
            st.session_state.ee_regions[rname]["adj"] = new_adj

# ── Tab: Pricing Table ──
with tabs[-2]:
    st.markdown('<p class="section-header">Pricing Table</p>', unsafe_allow_html=True)
    st.markdown(
        '<p class="section-sub">Build tiered pricing grids — adjust tier boundaries, base price, and discount %s</p>',
        unsafe_allow_html=True,
    )

    # ── Defaults (all zeros) ──
    if "pt_base_price" not in st.session_state:
        st.session_state.pt_base_price = 0.0
    if "pt_increment" not in st.session_state:
        st.session_state.pt_increment = 5000
    if "pt_tiers" not in st.session_state:
        st.session_state.pt_tiers = [
            {"low": 0, "high": 0, "pct": 100.0},
        ]

    # ── Controls row ──
    ctrl1, ctrl2 = st.columns(2)
    with ctrl1:
        base_price = st.number_input(
            "Starting Price Per User ($)",
            value=st.session_state.pt_base_price,
            step=1.0,
            min_value=0.0,
            key="pt_base_input",
            format="%.2f",
        )
        st.session_state.pt_base_price = base_price
    with ctrl2:
        increment = st.number_input(
            "Tier Increment (users)",
            value=st.session_state.pt_increment,
            step=500,
            min_value=1,
            key="pt_inc_input",
            help="When you add a tier, the High boundary auto-increases by this amount",
        )
        st.session_state.pt_increment = increment

    # ── Tier editor ──
    st.markdown("---")
    st.markdown("**Build Your Own Grid**")
    st.caption("Set employee ranges and the % of the prior tier's rate. Tier 1 uses the starting price per user.")

    tiers = st.session_state.pt_tiers
    delete_tier_idx: int | None = None
    recalc_from: int | None = None

    for i, tier in enumerate(tiers):
        cols = st.columns([3, 3, 3, 0.5])
        with cols[0]:
            new_low = st.number_input(
                "Employees Low" if i == 0 else f"Low {i}",
                value=tier["low"],
                step=1,
                min_value=0,
                key=f"pt_low_{i}",
                label_visibility="collapsed" if i > 0 else "visible",
            )
            if new_low != tier["low"]:
                tier["low"] = new_low
                # If tier 1 low changed, keep high >= low
                if tier["high"] < new_low:
                    tier["high"] = new_low
                recalc_from = i + 1
        with cols[1]:
            new_high = st.number_input(
                "Employees High" if i == 0 else f"High {i}",
                value=tier["high"],
                step=1,
                min_value=tier["low"],
                key=f"pt_high_{i}",
                label_visibility="collapsed" if i > 0 else "visible",
            )
            if new_high != tier["high"]:
                tier["high"] = new_high
                recalc_from = i + 1
        with cols[2]:
            if i == 0:
                st.text_input(
                    "% of Prior Rate", value="Base", key=f"pt_pct_d_{i}", disabled=True,
                    label_visibility="visible",
                )
            else:
                tier["pct"] = st.number_input(
                    f"% of Prior {i}",
                    value=tier["pct"],
                    step=1.0,
                    min_value=1.0,
                    max_value=100.0,
                    key=f"pt_pct_{i}",
                    format="%.1f",
                    label_visibility="collapsed",
                )
        with cols[3]:
            if i > 0:
                if st.button("✕", key=f"pt_del_{i}", type="secondary"):
                    delete_tier_idx = i

    # Cascade boundary changes downward
    if recalc_from is not None:
        for j in range(recalc_from, len(tiers)):
            prev_high = tiers[j - 1]["high"]
            tiers[j]["low"] = prev_high + 1
            tiers[j]["high"] = tiers[j]["low"] + increment - 1

    if delete_tier_idx is not None:
        tiers.pop(delete_tier_idx)
        # Re-cascade lows after deletion
        for j in range(1, len(tiers)):
            tiers[j]["low"] = tiers[j - 1]["high"] + 1
        st.rerun()

    if st.button("＋ Add Tier", key="pt_add_tier"):
        last_high = tiers[-1]["high"] if tiers else 0
        new_low = last_high + 1
        tiers.append({"low": new_low, "high": new_low + increment - 1, "pct": 97.0})
        st.rerun()

    # ── Compute pricing grid ──
    grid_rows: list[dict] = []
    prev_rate = 0.0

    for i, tier in enumerate(tiers):
        low = tier["low"]
        high = tier["high"]
        users_in_tier = max(high - low + 1, 0)

        if i == 0:
            rate_per_user = base_price
            tier_total = rate_per_user * users_in_tier
        else:
            rate_per_user = prev_rate * (tier["pct"] / 100.0)
            tier_total = users_in_tier * rate_per_user

        overall_total = (grid_rows[-1]["Overall Total"] + tier_total) if grid_rows else tier_total
        eff_rate = overall_total / high if high > 0 else 0

        grid_rows.append({
            "Employees Low": low,
            "Employees High": high,
            "Price / User": rate_per_user,
            "Tier Total": tier_total,
            "Overall Total": overall_total,
            "Effective Rate": eff_rate,
            "% of Prior": "Base" if i == 0 else f"{tier['pct']:.0f}%",
        })
        prev_rate = rate_per_user

    # ── Output Table 1: Grid ──
    st.markdown("---")

    if grid_rows:
        grid_df = pd.DataFrame(grid_rows)
        display_cols = ["Employees Low", "Employees High", "Price / User", "Tier Total", "Overall Total", "Effective Rate", "% of Prior"]
        st.dataframe(
            grid_df[display_cols].style.format({
                "Employees Low": "{:,.0f}",
                "Employees High": "{:,.0f}",
                "Price / User": "${:,.2f}",
                "Tier Total": "${:,.2f}",
                "Overall Total": "${:,.2f}",
                "Effective Rate": "${:,.2f}",
            }),
            use_container_width=True,
            hide_index=True,
        )

    # ── Output Table 2: Additional Users ──
    st.markdown("---")
    st.markdown("**Additional Users**")
    st.caption("Shows what a client pays when they grow into a new tier")

    if grid_rows:
        add_rows: list[dict] = []
        for i, tier in enumerate(tiers):
            low = tier["low"]
            high = tier["high"]
            label = f"Up to {high:,}" if i == 0 else f"{low:,} – {high:,}"
            overall = grid_rows[i]["Overall Total"]
            eff = grid_rows[i]["Effective Rate"]
            add_users = f"${grid_rows[i]['Price / User']:,.2f}"

            add_rows.append({
                "Eligible Users": label,
                "Additional User Rate": add_users,
                "Total": overall,
                "Effective Rate": eff,
            })

        add_df = pd.DataFrame(add_rows)
        st.dataframe(
            add_df.style.format({
                "Total": "${:,.2f}",
                "Effective Rate": "${:,.2f}",
            }),
            use_container_width=True,
            hide_index=True,
        )

# ── Tab: Export ──
with tabs[-1]:
    st.markdown('<p class="section-header">Export</p>', unsafe_allow_html=True)
    st.markdown(
        '<p class="section-sub">Download a CSV per scenario, ready to import into Google Sheets</p>',
        unsafe_allow_html=True,
    )

    deal_name = st.text_input(
        "Deal / Account Name",
        placeholder="e.g. Acme Corp",
        help="Used to name the downloaded files.",
    )

    slug = re.sub(r"[^a-zA-Z0-9]+", "_", deal_name.strip()).strip("_") if deal_name.strip() else "deal"
    today = date.today().isoformat()

    export_scenarios = [s for s in results if s["Schedule"]]

    if export_scenarios:
        xlsx_bytes = build_export_xlsx(results)
        filename = f"{slug}_{today}.xlsx"
        st.download_button(
            "⬇ Download Excel",
            data=xlsx_bytes,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    else:
        st.info("Configure at least one scenario to enable export.")
